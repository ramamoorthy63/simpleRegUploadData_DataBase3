
# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
from datetime import datetime
import pymysql  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
#wb = load_workbook('//home//hakeem//Desktop//excel.xlsx')
wb = load_workbook("Job_Placement_Data.excel") 
  
#database connection
connection = pymysql.connect(host="localhost",user="root",passwd="",database="Student")
cursor = connection.cursor()

# create the sheet object 
sheet = wb.active 
today=datetime.now()  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 40
    sheet.column_dimensions['K'].width = 50
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 10
    sheet.column_dimensions['N'].width = 10
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Mobile No"
    sheet.cell(row=1, column=4).value = "Alternate No"
    sheet.cell(row=1, column=5).value = "Email Id"
    sheet.cell(row=1, column=6).value = "Address"
    sheet.cell(row=1, column=7).value = "Course Interested"
    sheet.cell(row=1, column=8).value = "Batch Preffered"
    sheet.cell(row=1, column=9).value = "How You Came To Know Us"
    sheet.cell(row=1, column=10).value = "Are You Experience or Fresher"
    sheet.cell(row=1, column=11).value = "Contact Person From Besant Technologies"
    sheet.cell(row=1, column=12).value = "Counselor"
    sheet.cell(row=1, column=13).value = "Fees"
    sheet.cell(row=1, column=14).value = "Comments"
  
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the mobile_no_field box 
    mobile_no_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the alternate_no_field box 
    alternate_no_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the course_field box 
    course_field.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the batch_preferred_field box 
    batch_preferred_field.focus_set() 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the email_id_field box 
    email_id_field.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    address_field.focus_set() 
# Function to set focus 
def focus7(event):
    # set focus on the address_field box 
    date_field.focus_set()  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    mobile_no_field.delete(0, END) 
    alternate_no_field.delete(0, END) 
    course_field.delete(0, END) 
    batch_preferred_field.delete(0, END) 
    email_id_field.delete(0, END) 
    address_field.delete(0, END) 
    date_field.delete(0, END) 
  
def sub_option():
    print (regDetails.get())
    if regDetails.get() == 1:
    	temp = IntVar()    
    	Checkbutton(root, text="Demo", variable=temp).grid(row=11,column=1, sticky=N)
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
    print (regDetails.get())      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        mobile_no_field.get() == "" and
        alternate_no_field.get() == "" and
        course_field.get() == "" and
        batch_preferred_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == "" and 
        date_field.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = date_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = name_field.get()  
        sheet.cell(row=current_row + 1, column=3).value = mobile_no_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = alternate_no_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=6).value = address_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = course_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = batch_preferred_field.get() 
        
        sheet.cell(row=current_row + 1, column=9).value = how_you_came_to_know_field.get() 
        sheet.cell(row=current_row + 1, column=10).value =experience_or_fresher_field.get() 
        sheet.cell(row=current_row + 1, column=11).value = contact_person_field.get() 
        sheet.cell(row=current_row + 1, column=12).value = counselor_field.get() 
        sheet.cell(row=current_row + 1, column=13).value = fees_field.get() 
        sheet.cell(row=current_row + 1, column=14).value = comment_field.get() 

	# Data inserting into database Start
        if regDetails.get() == 1 and EnquiryDetails.get() == 1:
                insert_reg_record = "INSERT INTO Registration(Date,Form_Number,Name,Course,Phone_Number,Email_Id,Address,Status) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(date_field.get(),course_field.get(),name_field.get(),mobile_no_field.get(),batch_preferred_field.get(),email_id_field.get(),address_field.get(),alternate_no_field.get())
                insert_Enquiry_record = "INSERT INTO Enquiry(Date,Form_Number,Name,Course,Phone_Number,Email_Id,Address,Status) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(date_field.get(),course_field.get(),name_field.get(),mobile_no_field.get(),batch_preferred_field.get(),email_id_field.get(),address_field.get(),alternate_no_field.get())
        elif regDetails.get() == 1:
                insert_reg_record = "INSERT INTO Registration(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Interested,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees,Comments ) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(date_field.get(),name_field.get(),mobile_no_field.get(),alternate_no_field.get(),email_id_field.get(),address_field.get(),course_field.get(),batch_preferred_field.get(),how_you_came_to_know_field.get(),experience_or_fresher_field.get(),contact_person_field.get(),counselor_field.get(),fees_field.get(),comment_field.get())
        elif EnquiryDetails.get() == 1:
                insert_Enquiry_record = "INSERT INTO Enquiry(Date,Form_Number,Name,Course,Phone_Number,Email_Id,Address,Status) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(date_field.get(),course_field.get(),name_field.get(),mobile_no_field.get(),batch_preferred_field.get(),email_id_field.get(),address_field.get(),alternate_no_field.get())
        else :
                print ("Invalid Details")
        if regDetails.get() == 1:
           cursor.execute(insert_reg_record)
        if EnquiryDetails.get() == 1:
           cursor.execute(insert_Enquiry_record)
        if regDetails.get() == 1 or EnquiryDetails.get() == 1:
           connection.commit()
	# Data inserting into database End 
  
        # save the file 
        wb.save('excel.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light blue') 
  
    # set the title of GUI window 
    root.title("Besant Technologies") 
  
    # set the configuration of GUI window 
    root.geometry("640x480") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="Besant Technologies\nEnquiry Form", fg="Red") 
  
    # create a Date label 
    date = Label(root, text="Date:",bg="light blue" ) 
    
    # create a Name label 
    name = Label(root, text="Name:",bg="light blue")
    # create a Course label 
    mobile_no = Label(root, text="Mobile No:", bg="light blue") 
  
    # create a Semester label 
    alternate_no = Label(root, text="Alternate No:", bg="light blue") 
  
    # create a Form No. lable 
    email_id = Label(root, text="Email Id:", bg="light blue") 
  
    # create a Contact No. label 
    address = Label(root, text="Address:", bg="light blue") 
  
    # create a Email id label 
    course = Label(root, text="Course Interested:", bg="light blue") 
  
    # create a address label 
    batch_preferred= Label(root, text="Batch Preferred:", bg="light blue") 
    
    # create a address label 
    how_you_came_to_know= Label(root, text="How You Came To Know Us:", bg="light blue") 
    experience_or_fresher= Label(root, text="Are You Experience or Fresher:", bg="light blue")
    contact_person= Label(root, text="Contact Person From Besant Technologies:", bg="light blue") 
  
    counselor= Label(root, text="Counselor:", bg="light blue")
    fees= Label(root, text="Fees:", bg="light blue")
    comment= Label(root, text="Comment:", bg="light blue")
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    date.grid(row=1, column=0,sticky=W) 
    name.grid(row=2, column=0,sticky=W) 
    mobile_no.grid(row=3, column=0,sticky=W) 
    alternate_no.grid(row=4, column=0,sticky=W) 
    email_id.grid(row=5, column=0,sticky=W) 
    address.grid(row=6, column=0,sticky=W) 
    course.grid(row=7, column=0,sticky=W) 
    batch_preferred.grid(row=8, column=0,sticky=W) 
    how_you_came_to_know.grid(row=9,column=0,sticky=W)
    experience_or_fresher.grid(row=10,column=0,sticky=W)
    contact_person.grid(row=11,column=0,sticky=W)   
    counselor.grid(row=12,column=0,sticky=W)
    fees.grid(row=13,column=0,sticky=W)
    comment.grid(row=14,column=0,sticky=W)	

    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    mobile_no_field = Entry(root) 
    alternate_no_field = Entry(root) 
    course_field = Entry(root) 
    batch_preferred_field = Entry(root) 
    email_id_field = Entry(root) 
    address_field = Entry(root) 
    date_field = Entry(root) 
    how_you_came_to_know_field=Entry(root)
    experience_or_fresher_field=Entry(root)
    contact_person_field=Entry(root)
    counselor_field = Entry(root)
    fees_field = Entry(root)
    comment_field=Entry(root) 


    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    mobile_no_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    alternate_no_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    course_field.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    batch_preferred_field.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    email_id_field.bind("<Return>", focus6) 
    date_field.bind("<Return>", focus7) 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    date_field.grid(row=1, column=1,ipadx="80") 
    name_field.grid(row=2, column=1, ipadx="80") 
    mobile_no_field.grid(row=3, column=1, ipadx="80") 
    alternate_no_field.grid(row=4, column=1, ipadx="80") 
    email_id_field.grid(row=5, column=1, ipadx="80") 
    address_field.grid(row=6, column=1, ipadx="80") 
    course_field.grid(row=7, column=1, ipadx="80") 
    batch_preferred_field.grid(row=8, column=1, ipadx="80") 
    how_you_came_to_know_field.grid(row=9,column=1,ipadx="80")
    experience_or_fresher_field.grid(row=10,column=1,ipadx="80")
    contact_person_field.grid(row=11,column=1,ipadx="80")
    counselor_field.grid(row=12,column=1,ipadx="80")
    fees_field.grid(row=13,column=1,ipadx="80")
    comment_field.grid(row=14,column=1,ipadx="80")
    # call excel function 
    excel() 
    regDetails = IntVar()
    #Checkbutton(root, text="Registration", variable=regDetails).grid(row=9,column=1, sticky=N)
    Checkbutton(root, text="Registration", variable=regDetails,command=sub_option).grid(row=15,column=1, sticky=N)
    EnquiryDetails = IntVar()
    Checkbutton(root, text="Enquiry", variable=EnquiryDetails).grid(row=15,column=1, sticky=W)  

    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Green", command=insert) 
    submit.grid(row=16, column=1,sticky=W) 
  
    #Quit Button creating
    quite = Button(root, text="Quite", fg="Black",
                            bg="Red", command=root.destroy)
    quite.grid(row=16, column=1,sticky=N)

    # start the GUI 
    root.mainloop() 

